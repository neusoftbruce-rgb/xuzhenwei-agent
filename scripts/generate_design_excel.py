#!/usr/bin/env python3
"""生成设计书 v4.0 Excel 版本"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import yaml, os

wb = openpyxl.Workbook()

# ---- 样式 ----
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
SUB_FILL = PatternFill(start_color="E8E4FF", end_color="E8E4FF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="4ECB71", end_color="4ECB71", fill_type="solid")
GOLD_FILL = PatternFill(start_color="F0A050", end_color="F0A050", fill_type="solid")
TIPS_FILL = PatternFill(start_color="5B8DEF", end_color="5B8DEF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
WRAP = Alignment(wrap_text=True, vertical='top')

def style_header(ws, row, cols, fill=HEADER_FILL):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def style_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        cell.alignment = WRAP

def auto_width(ws, cols, max_w=50):
    for c in range(1, cols+1):
        ws.column_dimensions[get_column_letter(c)].width = min(max_w, 18 if c > 1 else 12)

# ============================================================
# Sheet 1: 设计概述
# ============================================================
ws1 = wb.active
ws1.title = "设计概述"
overview = [
    ["徐振伟智能体 — 系统设计书 v4.0", "", "", ""],
    ["版本", "4.0", "日期", "2026-06-17"],
    ["技术栈", "Spring Boot 3.4.7 + Spring AI 1.0.3 + DeepSeek API", "", ""],
    ["", "", "", ""],
    ["核心设计理念：双层独立叠加", "", "", ""],
    ["Layer 1: 思考深度", "🧠 深度思考 (默认)", "↔", "⚡ 快速模式"],
    ["Layer 2: 分析方式", "💬 自由对话", "↔", "🔧 技法分析 (106种)"],
    ["", "", "", ""],
    ["4种组合模式", "", "", ""],
    ["① 🧠+💬 深度自由", "DeepSeek Reasoner → 自然对话", "", ""],
    ["② 🧠+🔧 深度技法", "推理模型 → 技法工作流 → 精炼输出", "", ""],
    ["③ ⚡+💬 快速自由", "DeepSeek Chat → 直接回答", "", ""],
    ["④ ⚡+🔧 快速技法", "快模型 → 技法工作流(跳深度推理)", "", ""],
]
for i, row in enumerate(overview, 1):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)
auto_width(ws1, 6, 40)

# ============================================================
# Sheet 2: 56创意技法
# ============================================================
ws2 = wb.create_sheet("56创意技法")
headers2 = ["编号", "技法名称", "所属分类", "所属子类", "描述", "步骤数"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, 6)

# Read techniques.yml
with open(r"D:\myclaude\projects\xuzhenwei-agent\src\main\resources\techniques.yml", encoding='utf-8') as f:
    tech_data = yaml.safe_load(f)

CAT_CHAPTERS = {
    '快速产生灵感': [
        ('用AI特有力量思考', ['001','002','003','004','005','006']),
        ('用自由发想思考', ['007','008','009','010','011','012']),
        ('用逻辑发想思考', ['013','014','015','016','017']),
    ],
    '打磨完善创意': [
        ('让思考发展', ['018','019','020','021','022','023']),
        ('让思考具体化', ['024','025','026']),
        ('让思考验证', ['027','028','029','030','031','032']),
    ],
    '落地执行创意': [
        ('思考传播方式', ['033','034','035','036']),
        ('思考执行策略', ['037','038']),
    ],
    '找到思考切入点': [
        ('分析课题得启示', ['039','040','041','042','043']),
        ('分析烦恼得启示', ['044','045','046']),
        ('分析人得启示', ['047','048','049','050','051','052','053']),
        ('预测未来得启示', ['054','055','056']),
    ],
}

by_id = {t['id']: t for t in tech_data['techniques']}
row = 2
for cat in ['快速产生灵感','打磨完善创意','落地执行创意','找到思考切入点']:
    for sub_name, ids in CAT_CHAPTERS[cat]:
        for tid in ids:
            t = by_id.get(tid)
            if not t: continue
            ws2.cell(row=row, column=1, value=t['id'])
            ws2.cell(row=row, column=2, value=t['name'])
            ws2.cell(row=row, column=3, value=cat)
            ws2.cell(row=row, column=4, value=sub_name)
            ws2.cell(row=row, column=5, value=t.get('description',''))
            ws2.cell(row=row, column=6, value=t.get('stepCount',1))
            style_row(ws2, row, 6)
            row += 1

ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 50
ws2.auto_filter.ref = f"A1:F{row-1}"
ws2.freeze_panes = "A2"

# ============================================================
# Sheet 3: 50 AI独学TIPS
# ============================================================
ws3 = wb.create_sheet("50 AI独学TIPS")
headers3 = ["编号", "名称", "所属章节", "分类", "难度", "功能描述", "步骤数"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, 7, TIPS_FILL)

with open(r"D:\myclaude\projects\xuzhenwei-agent\src\main\resources\ai-self-study-tips.yml", encoding='utf-8') as f:
    tips_data = yaml.safe_load(f)

row = 2
for t in tips_data['tips']:
    ws3.cell(row=row, column=1, value=t['id'])
    ws3.cell(row=row, column=2, value=t['name'])
    ws3.cell(row=row, column=3, value=f"第{t.get('chapter','?')}章")
    ws3.cell(row=row, column=4, value=t['category'])
    ws3.cell(row=row, column=5, value=t.get('difficulty',''))
    ws3.cell(row=row, column=6, value=t.get('description',''))
    ws3.cell(row=row, column=7, value=t.get('stepCount',1))
    style_row(ws3, row, 7)
    row += 1

ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['F'].width = 55
ws3.auto_filter.ref = f"A1:G{row-1}"
ws3.freeze_panes = "A2"

# ============================================================
# Sheet 4: 15组交叉配对
# ============================================================
ws4 = wb.create_sheet("15组交叉配对")
headers4 = ["#", "56技法", "50TIPS配对", "互补逻辑"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, 4, GOLD_FILL)

cross_pairs = [
    (1, "001 跨界特征联想法", "TIPS-44 隐喻魔术师", "生成跨界创意后，为不同受众定制比喻"),
    (2, "005 虚拟专家会诊", "TIPS-11 AI陪练", "专家输出后AI陪练苏格拉底式追问"),
    (3, "010 跨界杂交法", "TIPS-35 概念图", "杂交概念→二维关系图可视化"),
    (4, "012 无限资源畅想", "TIPS-40 替代方案生成器", "理想版降级时爆出多种现实版方案"),
    (5, "015 10年倒推法", "TIPS-48 未来预测倒推", "业务倒推+个人倒推交叉验证"),
    (6, "017 SCAMPER九变", "TIPS-38 思维链(CoT)", "SCAMPER发散→CoT展开因果链"),
    (7, "018 盲区扫描法", "TIPS-42 认知偏差分析", "方案盲区+人的盲区，内外双重"),
    (8, "022 方案联姻法", "TIPS-43 灵感→框架", "联姻方案即时结构化"),
    (9, "025 6W3H策划案", "TIPS-12 Copilot深度分析", "策划维度→深度分析填充"),
    (10, "029 利益相关方预演", "TIPS-39 恶魔代言人", "外部视角+内部逻辑审查"),
    (11, "033 新闻发布稿法", "TIPS-41 AI投资人审查", "愿景故事+商业闭环验证"),
    (12, "035 第一批客户获取", "TIPS-25 策略性取舍", "获客渠道选择→取舍分析"),
    (13, "037 最小验证法POC", "TIPS-24 分级测试", "核心假设→三级难度验证"),
    (14, "040 障碍根因分析", "TIPS-16 预测绊脚石", "事后回溯+事前预判，闭环"),
    (15, "046 烦恼抽象化法", "TIPS-26 费曼技巧", "找到解法→费曼复述检验理解"),
]
for i, row_data in enumerate(cross_pairs, 2):
    for j, val in enumerate(row_data, 1):
        ws4.cell(row=i, column=j, value=val)
    style_row(ws4, i, 4)
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 55
ws4.freeze_panes = "A2"

# ============================================================
# Sheet 5: 5条农业混合工作流
# ============================================================
ws5 = wb.create_sheet("农业混合工作流")
headers5 = ["工作流", "步骤", "技法ID", "技法名称", "该步骤做什么"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, 5, GREEN_FILL)

workflows = [
    ("🌱 农技推广：发现翻译传播", [
        (1, "054", "最新趋势扫描法", "扫描农业前沿技术/政策趋势"),
        (2, "TIPS-44", "隐喻魔术师", "科研术语→农民能懂的比喻"),
        (3, "TIPS-07", "专业术语转通俗语言", "降到初中生能理解的水平"),
        (4, "036", "30秒视频脚本生成", "产出可传播的短视频脚本+PPT"),
    ]),
    ("📋 课程开发：策划验证迭代", [
        (1, "025", "6W3H策划案法", "课程框架：谁学/学什么/多久/多少钱"),
        (2, "029", "利益相关方预演法", "模拟学员/政府/讲师各方反应"),
        (3, "037", "最小验证法POC", "设计试听课验证课程需求"),
        (4, "027", "可行性打分法", "5维度量化评分"),
        (5, "TIPS-33", "周报PDCA", "课程上线后每周迭代改进"),
    ]),
    ("🏥 农场诊断：把脉开方跟诊", [
        (1, "005", "虚拟专家会诊", "种植/市场/文旅/财务多角色诊断"),
        (2, "040", "障碍根因分析法", "5个为什么深挖到根"),
        (3, "022", "方案联姻法", "多个建议整合为一个可执行方案"),
        (4, "038", "行动计划生成法", "WBS拆到每周每天做什么"),
        (5, "TIPS-19", "召唤计时员AI", "每日执行提醒和节奏管理"),
    ]),
    ("🏷️ 农业品牌：洞察定位内容", [
        (1, "047", "用户困扰调研法", "发现消费者对农产品的不满"),
        (2, "048", "不满清单法", "系统性梳理15+条不满"),
        (3, "051", "核心需求提炼法", "找到真正的购买驱动力"),
        (4, "023", "跨界商业模式借鉴", "借鉴成功消费品牌模式"),
        (5, "001", "跨界特征联想法", "从自然/生物特征找品牌差异化"),
    ]),
    ("🎮 学员养成：学练考用闭环", [
        (1, "TIPS-14", "定制学习计划", "AI根据学员背景出个性化路线图"),
        (2, "TIPS-24", "分级习题", "基础题→应用题→挑战题逐级检验"),
        (3, "TIPS-27", "与自己分身对话", "向AI讲解，检验是否真懂"),
        (4, "037", "最小验证法POC", "在自家农场实地验证所学"),
    ]),
]
row = 2
for wf_name, steps in workflows:
    for step_num, tid, tname, tdesc in steps:
        ws5.cell(row=row, column=1, value=wf_name)
        ws5.cell(row=row, column=2, value=f"第{step_num}步")
        ws5.cell(row=row, column=3, value=tid)
        ws5.cell(row=row, column=4, value=tname)
        ws5.cell(row=row, column=5, value=tdesc)
        style_row(ws5, row, 5)
        row += 1

ws5.column_dimensions['A'].width = 28
ws5.column_dimensions['D'].width = 25
ws5.column_dimensions['E'].width = 45
ws5.freeze_panes = "A2"

# ============================================================
# Sheet 6: 13配方套餐
# ============================================================
ws6 = wb.create_sheet("配方套餐(8+5)")
headers6 = ["类型", "配方名称", "技法链", "步骤数", "适用场景"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, 5)

recipes = [
    ("原8配方", "创业验证套餐", "002→026→034→031", 4, "创业/新业务/新项目"),
    ("原8配方", "内容创作套餐", "009→024→036", 3, "选题/内容/短视频/文案"),
    ("原8配方", "产品定价套餐", "034→050→027", 3, "定价/收费/价格策略"),
    ("原8配方", "客户获取套餐", "035→048→053", 3, "获客/引流/招生"),
    ("原8配方", "战略规划套餐", "005→015→030→038", 4, "战略/规划/转型"),
    ("原8配方", "提案打磨套餐", "025→031→033", 3, "提案/汇报/策划案"),
    ("原8配方", "农业品牌套餐", "001→005→028→036", 4, "农产品品牌"),
    ("原8配方", "深度诊断套餐", "040→044→045→030", 4, "诊断/根因分析"),
    ("🌾新增", "农技推广：发现翻译传播", "054→TIPS-44→TIPS-07→036", 4, "农业技术传播推广"),
    ("🌾新增", "课程开发：策划验证迭代", "025→029→037→027→TIPS-33", 5, "培训课程开发"),
    ("🌾新增", "农场诊断：把脉开方跟诊", "005→040→022→038→TIPS-19", 5, "农场经营诊断辅导"),
    ("🌾新增", "农业品牌：洞察定位内容", "047→048→051→023→001", 5, "农产品品牌打造"),
    ("🌾新增", "学员养成：学练考用闭环", "TIPS-14→TIPS-24→TIPS-27→037", 4, "培训后行为养成"),
]
for i, row_data in enumerate(recipes, 2):
    for j, val in enumerate(row_data, 1):
        ws6.cell(row=i, column=j, value=val)
    style_row(ws6, i, 5)
ws6.column_dimensions['B'].width = 28
ws6.column_dimensions['C'].width = 40
ws6.column_dimensions['E'].width = 30
ws6.freeze_panes = "A2"

# ============================================================
# Sheet 7: 实施路线图
# ============================================================
ws7 = wb.create_sheet("实施路线图")
headers7 = ["阶段", "内容", "涉及文件", "预计工时"]
for i, h in enumerate(headers7, 1):
    ws7.cell(row=1, column=i, value=h)
style_header(ws7, 1, 4, GOLD_FILL)

roadmap = [
    ("P1", "后端4模式分流", "AgentController + AgentEngine", "1h"),
    ("P2", "50TIPS注册到Registry", "TechniqueRegistry + ai-self-study-tips.yml", "1h"),
    ("P3", "混合推荐算法升级", "TechniqueRecommender", "1.5h"),
    ("P4", "前端双按钮+状态管理", "index.html (JS+CSS)", "2h"),
    ("P5", "侧边栏新增TIPS区域", "index.html (HTML+JS)", "1h"),
    ("P6", "5条混合工作流→配方", "AgentEngine (新增RecipeDef)", "1h"),
    ("P7", "5条新增技法→YAML", "techniques.yml (追加)", "1.5h"),
    ("P8", "联调测试", "全部", "1h"),
]
for i, row_data in enumerate(roadmap, 2):
    for j, val in enumerate(row_data, 1):
        ws7.cell(row=i, column=j, value=val)
    style_row(ws7, i, 4)
ws7.column_dimensions['B'].width = 30
ws7.column_dimensions['C'].width = 45
ws7.freeze_panes = "A2"

# ============================================================
# Sheet 8: 系统盲区
# ============================================================
ws8 = wb.create_sheet("系统盲区与待改进")
headers8 = ["盲区", "描述", "农业场景后果", "建议方案"]
for i, h in enumerate(headers8, 1):
    ws8.cell(row=1, column=i, value=h)
style_header(ws8, 1, 4)

gaps = [
    ("群体共创引导", "所有技法均为单用户+AI模式", "无法在15-30人培训班集体使用", "设计多人协作技法，AI做中立主持人"),
    ("物理空间联动", "无实景互动设计", "不能站在大棚里用AI分析当前作物", "增加图像识别+实地输入接口"),
    ("长期行为改变", "有机制化但偏AI使用机制", "学完回去不执行", "增加21天习惯养成+accountability小组"),
    ("跨语言适配", "纯中文场景", "无法检索日语六次产业化原始文献", "增加多语言prompt模板"),
    ("数据采集闭环", "输出是终点不是起点", "学员反馈无法汇聚成改进报告", "增加反馈收集+自动聚合分析"),
]
for i, row_data in enumerate(gaps, 2):
    for j, val in enumerate(row_data, 1):
        ws8.cell(row=i, column=j, value=val)
    style_row(ws8, i, 4)
ws8.column_dimensions['B'].width = 30
ws8.column_dimensions['C'].width = 35
ws8.column_dimensions['D'].width = 35
ws8.freeze_panes = "A2"

# ============================================================
# 保存
# ============================================================
output_path = r"D:\myclaude\projects\xuzhenwei-agent\设计书_v4.0.xlsx"
wb.save(output_path)
print(f"Excel saved: {output_path}")
print(f"Sheet count: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
